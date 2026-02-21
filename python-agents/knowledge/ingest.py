"""Knowledge ingestion engine.

Handles two ingestion paths:
1. File uploads (PDF, DOCX, TXT, MD, CSV, XLSX)
2. URL scraping (public pages only)

All ingested content is chunked and returned for storage.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

import httpx
from bs4 import BeautifulSoup

from config import (
    KB_ALLOWED_FILE_TYPES,
    KB_MAX_CONTENT_CHARS,
    KB_MAX_FILE_SIZE,
)
from knowledge.chunker import chunk_text


# ---------------------------------------------------------------------------
# File ingestion
# ---------------------------------------------------------------------------


def extract_text_from_file(content: bytes, filename: str) -> str:
    """Extract plain text from a file based on its extension."""
    ext = Path(filename).suffix.lower()

    if ext not in KB_ALLOWED_FILE_TYPES:
        raise ValueError(
            f"Unsupported file type: {ext}. Allowed: {', '.join(sorted(KB_ALLOWED_FILE_TYPES))}"
        )

    if len(content) > KB_MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {len(content)} bytes. Maximum: {KB_MAX_FILE_SIZE // (1024*1024)}MB"
        )

    if ext == ".pdf":
        return _extract_pdf(content)
    elif ext == ".docx":
        return _extract_docx(content)
    elif ext in (".txt", ".md"):
        return content.decode("utf-8", errors="replace")
    elif ext == ".csv":
        return _extract_csv(content)
    elif ext == ".xlsx":
        return _extract_xlsx(content)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _extract_pdf(content: bytes) -> str:
    """Extract text from PDF using pdfplumber."""
    import pdfplumber

    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n\n".join(text_parts)


def _extract_docx(content: bytes) -> str:
    """Extract text from DOCX using python-docx."""
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_csv(content: bytes) -> str:
    """Convert CSV to structured text."""
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return ""

    # Use first row as headers
    headers = rows[0]
    lines: list[str] = []
    for row in rows[1:]:
        pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
        if pairs:
            lines.append(" | ".join(pairs))
    return "\n".join(lines)


def _extract_xlsx(content: bytes) -> str:
    """Convert XLSX to structured text."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines: list[str] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        lines.append(f"--- Sheet: {sheet} ---")
        headers = [str(h or "") for h in rows[0]]
        for row in rows[1:]:
            pairs = [
                f"{h}: {v}"
                for h, v in zip(headers, row)
                if v is not None and str(v).strip()
            ]
            if pairs:
                lines.append(" | ".join(pairs))

    return "\n".join(lines)


def ingest_file(content: bytes, filename: str) -> dict[str, Any]:
    """Full file ingestion pipeline: extract → truncate → chunk.

    Returns: {
        "source_name": str,
        "content": str,
        "content_chunks": list[str],
        "file_type": str,
        "file_size": int,
        "char_count": int,
        "chunk_count": int,
    }
    """
    text = extract_text_from_file(content, filename)

    # Truncate if needed
    if len(text) > KB_MAX_CONTENT_CHARS:
        text = text[:KB_MAX_CONTENT_CHARS] + "\n\n[Content truncated at 50,000 characters]"

    chunks = chunk_text(text)
    ext = Path(filename).suffix.lower()

    return {
        "source_name": filename,
        "content": text,
        "content_chunks": chunks,
        "file_type": ext.lstrip("."),
        "file_size": len(content),
        "char_count": len(text),
        "chunk_count": len(chunks),
    }


# ---------------------------------------------------------------------------
# URL scraping
# ---------------------------------------------------------------------------


async def scrape_url(url: str) -> dict[str, Any]:
    """Scrape a public URL and extract main content.

    Returns: {
        "source_name": str (page title),
        "source_url": str,
        "content": str,
        "content_chunks": list[str],
        "char_count": int,
        "chunk_count": int,
        "domain": str,
    }

    Raises ValueError if URL is not reachable or requires authentication.
    """
    # Validate URL format
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    try:
        async with httpx.AsyncClient(
            timeout=30,
            follow_redirects=True,
            headers={
                "User-Agent": "Mozilla/5.0 (Virtual PPO Knowledge Bot)"
            },
        ) as client:
            resp = await client.get(url)
    except httpx.RequestError as exc:
        raise ValueError(f"Could not reach URL: {exc}")

    # Check for auth walls
    if resp.status_code in (401, 403):
        raise ValueError(
            f"URL requires authentication (HTTP {resp.status_code}). "
            "Only public pages without login are supported."
        )

    if resp.status_code >= 400:
        raise ValueError(f"URL returned error: HTTP {resp.status_code}")

    # Detect login redirects
    final_url = str(resp.url)
    if any(kw in final_url.lower() for kw in ["login", "signin", "auth", "sso"]):
        raise ValueError(
            "URL redirected to a login page. Only public pages without authentication are supported."
        )

    # Parse HTML
    content_type = resp.headers.get("content-type", "")
    if "html" not in content_type.lower() and "text" not in content_type.lower():
        raise ValueError(f"URL did not return HTML content (got: {content_type})")

    soup = BeautifulSoup(resp.text, "html.parser")

    # Extract title
    title_tag = soup.find("title")
    title = title_tag.get_text(strip=True) if title_tag else url

    # Remove non-content elements
    for tag in soup(["script", "style", "nav", "footer", "header", "aside", "noscript", "iframe"]):
        tag.decompose()

    # Try to find main content
    main_content = (
        soup.find("main")
        or soup.find("article")
        or soup.find("div", {"role": "main"})
        or soup.find("div", class_=re.compile(r"content|main|article|body", re.I))
        or soup.body
    )

    if main_content is None:
        raise ValueError("Could not extract content from URL")

    text = main_content.get_text(separator="\n", strip=True)

    # Clean up whitespace
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r" {2,}", " ", text)

    # Truncate
    if len(text) > KB_MAX_CONTENT_CHARS:
        text = text[:KB_MAX_CONTENT_CHARS] + "\n\n[Content truncated at 50,000 characters]"

    chunks = chunk_text(text)

    # Extract domain
    from urllib.parse import urlparse
    domain = urlparse(url).netloc

    return {
        "source_name": title[:200],
        "source_url": url,
        "content": text,
        "content_chunks": chunks,
        "char_count": len(text),
        "chunk_count": len(chunks),
        "domain": domain,
    }
